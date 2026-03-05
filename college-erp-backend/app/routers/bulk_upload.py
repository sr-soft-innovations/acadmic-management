"""Bulk Excel/CSV upload for all main tables. First row = headers, then data rows."""
import csv
import re
import secrets
from io import BytesIO, StringIO

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File

from app.db import read_json, write_json, next_id
from app.routers.auth import _get_current_user, _hash_password, _validate_password_policy

router = APIRouter()

ALLOWED_ENTITIES = [
    "students",
    "staff",
    "courses",
    "timetable",
    "attendance",
    "fees",
    "exams",
    "subject_faculty",
    "semesters",
    "holidays",
    "users",
]


def _normalize_header(h: str) -> str:
    if h is None:
        return ""
    s = str(h).strip().lower()
    s = re.sub(r"\s+", "_", s)
    return s


def _cell_value(cell) -> str:
    if cell is None:
        return ""
    v = getattr(cell, "value", None)
    if v is None:
        return ""
    return str(v).strip()


def parse_excel(file_content: bytes) -> list[dict]:
    """Parse first sheet: row 1 = headers, rest = data. Returns list of dicts (header -> value)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. pip install openpyxl")
    wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        return []
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row or 0, values_only=False))
    if not rows:
        return []
    headers = [_normalize_header(_cell_value(c)) for c in rows[0]]
    out = []
    for row in rows[1:]:
        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = _cell_value(cell)
        if any(v for v in row_dict.values()):
            out.append(row_dict)
    return out


def parse_csv(file_content: bytes) -> list[dict]:
    """Parse CSV: first row = headers, rest = data. Returns list of dicts (normalized header -> value)."""
    try:
        text = file_content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = file_content.decode("latin-1")
    reader = csv.DictReader(StringIO(text))
    out = []
    for row in reader:
        normalized = {_normalize_header(k): str(v).strip() if v else "" for k, v in row.items() if k}
        if any(v for v in normalized.values()):
            out.append(normalized)
    return out


def _float(val, default=0):
    if val is None or val == "":
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _int(val, default=0):
    if val is None or val == "":
        return default
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def _str(val, default=""):
    if val is None:
        return default
    return str(val).strip() or default


def _process_students(rows: list[dict]) -> tuple[int, int, list[dict]]:
    created, updated, errors = 0, 0, []
    students = read_json("students")
    existing_roll = {str(s.get("roll_no", "")).strip(): s for s in students if (s.get("roll_no") or "").strip()}
    for idx, row in enumerate(rows):
        try:
            name = _str(row.get("name") or row.get("student_name"))
            if not name:
                errors.append({"row": idx + 2, "message": "name is required"})
                continue
            roll_no = _str(row.get("roll_no") or row.get("roll_number"))
            existing = existing_roll.get(roll_no) if roll_no else None
            rec = {
                "id": next_id("students") if not existing else existing["id"],
                "name": name,
                "roll_no": roll_no,
                "course": _str(row.get("course")),
                "semester": _str(row.get("semester")),
                "email": _str(row.get("email")),
                "phone": _str(row.get("phone")),
                "date_of_birth": _str(row.get("date_of_birth") or row.get("dob")),
                "gender": _str(row.get("gender")),
                "address": _str(row.get("address")),
                "guardian_name": _str(row.get("guardian_name")),
                "guardian_phone": _str(row.get("guardian_phone")),
                "category": _str(row.get("category")),
                "batch": _str(row.get("batch")),
                "section": _str(row.get("section")),
                "admission_date": _str(row.get("admission_date")),
                "academic_year": _str(row.get("academic_year")),
                "photo_filename": "",
                "is_alumni": False,
                "alumni_year": "",
            }
            if existing:
                for k, v in rec.items():
                    if k != "id" and v:
                        existing[k] = v
                updated += 1
            else:
                existing_roll[roll_no] = rec
                students.append(rec)
                created += 1
        except Exception as e:
            errors.append({"row": idx + 2, "message": str(e)})
    if created or updated:
        write_json("students", students)
    return created, updated, errors


def _process_staff(rows: list[dict]) -> tuple[int, int, list[dict]]:
    created, updated, errors = 0, 0, []
    staff_list = read_json("staff")
    existing_email = {str(s.get("email", "")).strip().lower(): s for s in staff_list if (s.get("email") or "").strip()}
    for idx, row in enumerate(rows):
        try:
            name = _str(row.get("name") or row.get("staff_name"))
            if not name:
                errors.append({"row": idx + 2, "message": "name is required"})
                continue
            email = _str(row.get("email"))
            existing = existing_email.get(email.lower()) if email else None
            rec = {
                "id": next_id("staff") if not existing else existing["id"],
                "name": name,
                "designation": _str(row.get("designation")),
                "department": _str(row.get("department")),
                "role": _str(row.get("role")) or "Faculty",
                "email": email,
                "phone": _str(row.get("phone")),
                "date_of_joining": _str(row.get("date_of_joining") or row.get("doj")),
                "is_guest_faculty": (row.get("is_guest_faculty") or row.get("guest") or "").lower() in ("1", "yes", "true"),
                "is_substitute": (row.get("is_substitute") or row.get("substitute") or "").lower() in ("1", "yes", "true"),
                "substitute_for_staff_id": _str(row.get("substitute_for_staff_id")),
                "qualifications": [],
                "experience": [],
            }
            if existing:
                for k, v in rec.items():
                    if k != "id" and v is not None and v != "":
                        existing[k] = v
                updated += 1
            else:
                existing_email[email.lower()] = rec
                staff_list.append(rec)
                created += 1
        except Exception as e:
            errors.append({"row": idx + 2, "message": str(e)})
    if created or updated:
        write_json("staff", staff_list)
    return created, updated, errors


def _process_courses(rows: list[dict]) -> tuple[int, int, list[dict]]:
    created, updated, errors = 0, 0, []
    courses = read_json("courses")
    existing_code = {str(c.get("code", "")).strip().lower(): c for c in courses if (c.get("code") or "").strip()}
    for idx, row in enumerate(rows):
        try:
            name = _str(row.get("name") or row.get("course_name"))
            if not name:
                errors.append({"row": idx + 2, "message": "name is required"})
                continue
            code = _str(row.get("code"))
            existing = existing_code.get(code.lower()) if code else existing_code.get(name.lower())
            type_val = (_str(row.get("type")) or "theory").lower()
            if type_val not in ("theory", "practical", "elective"):
                type_val = "theory"
            rec = {
                "id": next_id("courses") if not existing else existing["id"],
                "name": name,
                "code": code,
                "department": _str(row.get("department")),
                "semester": _str(row.get("semester")),
                "credits": _float(row.get("credits"), 0),
                "type": type_val,
                "hours_per_week": _float(row.get("hours_per_week") or row.get("hours"), 0),
            }
            if existing:
                for k, v in rec.items():
                    if k != "id":
                        existing[k] = v
                updated += 1
            else:
                courses.append(rec)
                if code:
                    existing_code[code.lower()] = rec
                created += 1
        except Exception as e:
            errors.append({"row": idx + 2, "message": str(e)})
    if created or updated:
        write_json("courses", courses)
    return created, updated, errors


def _process_timetable(rows: list[dict]) -> tuple[int, int, list[dict]]:
    created, updated, errors = 0, 0, []
    slots = read_json("timetable")
    subjects = read_json("courses")
    staff_list = read_json("staff")
    sub_ids = {str(s.get("name", "")).strip().lower(): str(s.get("id")) for s in subjects}
    staff_ids = {str(s.get("name", "")).strip().lower(): str(s.get("id")) for s in staff_list}
    for idx, row in enumerate(rows):
        try:
            day = _int(row.get("day_of_week") or row.get("day"), 1)
            if day < 1 or day > 6:
                errors.append({"row": idx + 2, "message": "day_of_week must be 1-6"})
                continue
            sub_name = _str(row.get("subject") or row.get("subject_name"))
            subject_id = _str(row.get("subject_id")) or (sub_ids.get(sub_name.lower()) if sub_name else "") or ""
            st_name = _str(row.get("staff") or row.get("staff_name"))
            staff_id = _str(row.get("staff_id")) or (staff_ids.get(st_name.lower()) if st_name else "") or ""
            if not subject_id or not staff_id:
                errors.append({"row": idx + 2, "message": "subject and staff required"})
                continue
            slot = {
                "id": next_id("timetable"),
                "day_of_week": day,
                "slot_start": _str(row.get("slot_start") or row.get("start")) or "09:00",
                "slot_end": _str(row.get("slot_end") or row.get("end")) or "10:00",
                "subject_id": subject_id,
                "staff_id": staff_id,
                "room": _str(row.get("room")),
                "program": _str(row.get("program")),
                "semester": _str(row.get("semester")),
                "slot_type": (_str(row.get("slot_type")) or "theory").lower() or "theory",
                "lab_batch": _str(row.get("lab_batch")),
                "is_extra": False,
                "slot_date": "",
            }
            slots.append(slot)
            created += 1
        except Exception as e:
            errors.append({"row": idx + 2, "message": str(e)})
    if created:
        write_json("timetable", slots)
    return created, updated, errors


def _process_attendance(rows: list[dict]) -> tuple[int, int, list[dict]]:
    created, updated, errors = 0, 0, []
    items = read_json("attendance")
    students = {str(s.get("id")) for s in read_json("students")}
    for idx, row in enumerate(rows):
        try:
            student_id = _str(row.get("student_id"))
            date_str = _str(row.get("date"))[:10]
            if not student_id or not date_str:
                errors.append({"row": idx + 2, "message": "student_id and date required"})
                continue
            if student_id not in students:
                errors.append({"row": idx + 2, "message": f"student_id {student_id} not found"})
                continue
            status = (_str(row.get("status")) or "present").lower() or "present"
            items.append({
                "id": next_id("attendance"),
                "student_id": student_id,
                "date": date_str,
                "status": status,
                "subject_id": _str(row.get("subject_id")),
                "course_id": _str(row.get("subject_id")),
                "source": "bulk_excel",
                "period_number": _int(row.get("period_number") or row.get("period")),
                "slot_type": (_str(row.get("slot_type")) or "theory").lower() or "theory",
            })
            created += 1
        except Exception as e:
            errors.append({"row": idx + 2, "message": str(e)})
    if created:
        write_json("attendance", items)
    return created, updated, errors


def _process_fees(rows: list[dict]) -> tuple[int, int, list[dict]]:
    from datetime import datetime, timezone
    created, updated, errors = 0, 0, []
    collections = read_json("fee_collections")
    students = {str(s.get("id")): s for s in read_json("students")}
    for idx, row in enumerate(rows):
        try:
            student_id = _str(row.get("student_id"))
            amount = _float(row.get("amount"))
            if not student_id or amount <= 0:
                errors.append({"row": idx + 2, "message": "student_id and amount > 0 required"})
                continue
            if student_id not in students:
                errors.append({"row": idx + 2, "message": f"student_id {student_id} not found"})
                continue
            s = students[student_id]
            paid_at = datetime.now(timezone.utc).isoformat()
            collections.append({
                "id": next_id("fee_collections"),
                "student_id": student_id,
                "student_name": s.get("name", ""),
                "roll_no": s.get("roll_no", ""),
                "course": _str(row.get("course")) or s.get("course", ""),
                "semester": _str(row.get("semester")) or s.get("semester", ""),
                "amount": amount,
                "fee_type": (_str(row.get("fee_type")) or "tuition").lower() or "tuition",
                "receipt_no": _str(row.get("receipt_no") or row.get("receipt")),
                "remarks": _str(row.get("remarks")),
                "paid_at": paid_at,
                "created_at": paid_at,
            })
            created += 1
        except Exception as e:
            errors.append({"row": idx + 2, "message": str(e)})
    if created:
        write_json("fee_collections", collections)
    return created, updated, errors


def _process_exams(rows: list[dict]) -> tuple[int, int, list[dict]]:
    created, updated, errors = 0, 0, []
    exams = read_json("exams")
    for idx, row in enumerate(rows):
        try:
            title = _str(row.get("title") or row.get("exam_name"))
            if not title:
                errors.append({"row": idx + 2, "message": "title is required"})
                continue
            exams.append({
                "id": next_id("exams"),
                "title": title,
                "course": _str(row.get("course")),
                "subject": _str(row.get("subject")),
                "subject_id": _str(row.get("subject_id")),
                "exam_type": (_str(row.get("exam_type")) or "internal").lower() or "internal",
                "exam_date": _str(row.get("exam_date") or row.get("date"))[:10],
                "start_time": _str(row.get("start_time") or row.get("time")),
                "duration_minutes": _int(row.get("duration_minutes") or row.get("duration"), 0),
                "room": _str(row.get("room")),
            })
            created += 1
        except Exception as e:
            errors.append({"row": idx + 2, "message": str(e)})
    if created:
        write_json("exams", exams)
    return created, updated, errors


def _process_subject_faculty(rows: list[dict]) -> tuple[int, int, list[dict]]:
    created, updated, errors = 0, 0, []
    items = read_json("subject_faculty")
    subjects = {str(s.get("id")) for s in read_json("courses")}
    staff_list = {str(s.get("id")) for s in read_json("staff")}
    seen = {(str(m.get("subject_id")), str(m.get("staff_id"))) for m in items}
    for idx, row in enumerate(rows):
        try:
            subject_id = _str(row.get("subject_id"))
            staff_id = _str(row.get("staff_id"))
            if not subject_id or not staff_id:
                errors.append({"row": idx + 2, "message": "subject_id and staff_id required"})
                continue
            if subject_id not in subjects or staff_id not in staff_list:
                errors.append({"row": idx + 2, "message": "subject_id or staff_id not found"})
                continue
            if (subject_id, staff_id) in seen:
                errors.append({"row": idx + 2, "message": "mapping already exists"})
                continue
            items.append({"id": next_id("subject_faculty"), "subject_id": subject_id, "staff_id": staff_id})
            seen.add((subject_id, staff_id))
            created += 1
        except Exception as e:
            errors.append({"row": idx + 2, "message": str(e)})
    if created:
        write_json("subject_faculty", items)
    return created, updated, errors


def _process_semesters(rows: list[dict]) -> tuple[int, int, list[dict]]:
    created, updated, errors = 0, 0, []
    items = read_json("semesters")
    for idx, row in enumerate(rows):
        try:
            program = _str(row.get("program"))
            sem_num = _int(row.get("semester_number") or row.get("semester"), 1)
            if not program:
                errors.append({"row": idx + 2, "message": "program is required"})
                continue
            name = _str(row.get("name")) or f"Semester {sem_num}"
            items.append({
                "id": next_id("semesters"),
                "program": program,
                "semester_number": sem_num,
                "name": name,
            })
            created += 1
        except Exception as e:
            errors.append({"row": idx + 2, "message": str(e)})
    if created:
        write_json("semesters", items)
    return created, updated, errors


def _process_users(rows: list[dict]) -> tuple[int, int, list[dict]]:
    created, updated, errors = 0, 0, []
    users = read_json("users")
    existing_username = {(u.get("username") or "").strip().lower(): u for u in users if (u.get("username") or "").strip()}
    for idx, row in enumerate(rows):
        try:
            username = _str(row.get("username"))
            if not username:
                errors.append({"row": idx + 2, "message": "username is required"})
                continue
            name = _str(row.get("name") or row.get("full_name"))
            role = _str(row.get("role")) or "staff"
            email = _str(row.get("email"))
            phone = _str(row.get("phone"))
            password = _str(row.get("password"))
            if not password:
                password = secrets.token_urlsafe(12)
            else:
                errs = _validate_password_policy(password)
                if errs:
                    errors.append({"row": idx + 2, "message": "; ".join(errs)})
                    continue
            existing = existing_username.get(username.lower())
            if existing:
                errors.append({"row": idx + 2, "message": f"username {username} already exists"})
                continue
            new_id = next_id("users")
            rec = {
                "id": new_id,
                "username": username,
                "password": _hash_password(password),
                "name": name or username,
                "email": email,
                "phone": phone,
                "role": role,
                "photo_filename": "",
                "is_enabled": True,
                "failed_attempts": 0,
                "locked_until": None,
                "password_changed_at": None,
                "password_history": [],
            }
            if role == "student":
                rec["linked_student_id"] = _str(row.get("linked_student_id"))
            elif role == "parent":
                linked = _str(row.get("linked_student_ids") or row.get("linked_student_id"))
                rec["linked_student_ids"] = [s.strip() for s in linked.split(",") if s.strip()]
            elif role == "hod":
                rec["department"] = _str(row.get("department"))
            elif role in ("faculty", "staff"):
                rec["linked_staff_id"] = _str(row.get("linked_staff_id"))
            users.append(rec)
            existing_username[username.lower()] = rec
            created += 1
        except Exception as e:
            errors.append({"row": idx + 2, "message": str(e)})
    if created:
        write_json("users", users)
    return created, updated, errors


def _process_holidays(rows: list[dict]) -> tuple[int, int, list[dict]]:
    created, updated, errors = 0, 0, []
    items = read_json("timetable_holidays")
    existing_dates = {(h.get("date") or "")[:10] for h in items}
    for idx, row in enumerate(rows):
        try:
            date_str = _str(row.get("date"))[:10]
            if not date_str or len(date_str) < 10:
                errors.append({"row": idx + 2, "message": "date (YYYY-MM-DD) required"})
                continue
            if date_str in existing_dates:
                errors.append({"row": idx + 2, "message": "holiday already exists"})
                continue
            items.append({
                "id": next_id("timetable_holidays"),
                "date": date_str,
                "name": _str(row.get("name") or row.get("holiday_name")),
            })
            existing_dates.add(date_str)
            created += 1
        except Exception as e:
            errors.append({"row": idx + 2, "message": str(e)})
    if created:
        write_json("timetable_holidays", items)
    return created, updated, errors


HANDLERS = {
    "students": _process_students,
    "staff": _process_staff,
    "courses": _process_courses,
    "timetable": _process_timetable,
    "attendance": _process_attendance,
    "fees": _process_fees,
    "exams": _process_exams,
    "subject_faculty": _process_subject_faculty,
    "semesters": _process_semesters,
    "holidays": _process_holidays,
    "users": _process_users,
}


@router.get("/templates/{entity}")
def get_template_columns(entity: str, current: dict = Depends(_get_current_user)):
    """Return expected Excel column headers for each entity (for download template)."""
    if entity not in ALLOWED_ENTITIES:
        raise HTTPException(status_code=400, detail=f"Unknown entity. Allowed: {ALLOWED_ENTITIES}")
    templates = {
        "students": ["name", "roll_no", "course", "semester", "email", "phone", "date_of_birth", "gender", "address", "guardian_name", "guardian_phone", "category", "batch", "section", "admission_date", "academic_year"],
        "staff": ["name", "designation", "department", "role", "email", "phone", "date_of_joining", "is_guest_faculty", "is_substitute"],
        "courses": ["name", "code", "department", "semester", "credits", "type", "hours_per_week"],
        "timetable": ["day_of_week", "slot_start", "slot_end", "subject_id", "staff_id", "room", "program", "semester", "slot_type", "lab_batch"],
        "attendance": ["student_id", "date", "status", "subject_id", "period_number", "slot_type"],
        "fees": ["student_id", "amount", "fee_type", "receipt_no", "course", "semester", "remarks"],
        "exams": ["title", "course", "subject", "subject_id", "exam_type", "exam_date", "start_time", "duration_minutes", "room"],
        "subject_faculty": ["subject_id", "staff_id"],
        "semesters": ["program", "semester_number", "name"],
        "holidays": ["date", "name"],
        "users": ["username", "name", "role", "email", "phone", "password", "linked_student_id", "linked_student_ids", "linked_staff_id", "department"],
    }
    return {"entity": entity, "columns": templates.get(entity, [])}


@router.post("/{entity}")
async def upload_excel(
    entity: str,
    file: UploadFile = File(...),
    current: dict = Depends(_get_current_user),
):
    """Upload Excel (.xlsx) or CSV (.csv) for bulk insert. First row = headers, rest = data. Entity: students, staff, courses, timetable, attendance, fees, exams, subject_faculty, semesters, holidays."""
    if entity not in ALLOWED_ENTITIES:
        raise HTTPException(status_code=400, detail=f"Unknown entity. Allowed: {ALLOWED_ENTITIES}")
    fn = (file.filename or "").lower()
    if not fn.endswith(".xlsx") and not fn.endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only .xlsx or .csv file allowed")
    content = await file.read()
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 10MB)")
    try:
        rows = parse_csv(content) if fn.endswith(".csv") else parse_excel(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid file: {e}")
    if not rows:
        return {"created": 0, "updated": 0, "errors": [], "message": "No data rows found (first row should be headers)"}
    handler = HANDLERS[entity]
    created, updated, errors = handler(rows)
    return {
        "created": created,
        "updated": updated,
        "errors": errors[:100],
        "total_rows": len(rows),
        "message": f"Processed: {created} created, {updated} updated, {len(errors)} errors",
    }
